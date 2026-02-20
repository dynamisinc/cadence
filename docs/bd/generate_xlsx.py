"""Generate Cadence_Feature_Matrix.xlsx from the feature matrix data."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Feature Matrix'

# -- Styles --
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
category_font = Font(name='Calibri', size=11, bold=True, color='1A3C6E')
category_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
normal_font = Font(name='Calibri', size=10)
alt_fill = PatternFill(start_color='F2F6FC', end_color='F2F6FC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')
center_alignment = Alignment(horizontal='center', vertical='top')

# -- Data --
features = [
    # (category, number, feature_name, description)
    ('HSEEP Compliance & Framework Support', 1, 'HSEEP 2020 Doctrinal Compliance',
     'Full alignment with HSEEP 2020 doctrine including exercise types (TTX, FE, FSE, CAX), participant roles, evaluation methodology (P/S/M/U ratings), and documentation standards.'),
    ('', 2, 'FEMA PrepToolkit Inject Status Workflow',
     'Implements the FEMA PrepToolkit inject lifecycle: Draft, Submitted, Approved, Synchronized, Released, Complete, Deferred, and Obsolete statuses with enforced transition rules.'),
    ('', 3, 'Exercise Evaluation Guide (EEG)',
     'Structured evaluation framework implementing the HSEEP chain: Capability Targets with measurable thresholds, Critical Tasks, inject-to-task linkage, and EEG entries with mandatory P/S/M/U ratings.'),
    ('', 4, 'SMART Objective Management',
     'Define and manage HSEEP-compliant SMART objectives per exercise with many-to-many linkage to injects for coverage tracking and evaluation alignment.'),
    ('', 5, 'Nine HSEEP Participant Roles',
     'All HSEEP-defined roles with exercise-scoped permissions: Exercise Director, Controller, Evaluator, Observer, Player, Simulator, Facilitator, Safety Officer, and Trusted Agent.'),
    ('', 6, 'Cross-Domain Framework Support',
     'Configurable status workflows and terminology for DoD/JTS, NATO, UK Cabinet Office, Australian AIIMS, NIST/MITRE cybersecurity, CMS/Joint Commission healthcare, FFIEC/FINRA financial, and ISO 22301/BCI frameworks.'),
    ('', 7, 'P/S/M/U Performance Rating',
     'HSEEP-standard performance rating scale applied to both ad-hoc observations and structured EEG entries: Performed, Some Challenges, Major Challenges, Unable to Perform.'),

    ('Exercise Design & Planning', 8, 'Exercise CRUD',
     'Create, view, edit, and archive exercises with support for all HSEEP exercise types, automatic MSEL creation, and practice mode for training exercises excluded from production metrics.'),
    ('', 9, 'Exercise Status Workflow',
     'HSEEP-compliant lifecycle management through Draft, Active, Paused, Completed, and Archived states with validation rules, confirmation dialogs, and full audit trail.'),
    ('', 10, 'Dual-Time Tracking',
     'Native support for both wall clock (scheduled delivery) and scenario time (story time) on every inject, enabling multi-day scenario compression into shorter exercise windows.'),
    ('', 11, 'MSEL Authoring',
     'Full inject management with auto-sequential numbering, all standard MSEL fields (title, description, sender, recipient, delivery method, expected action, controller notes), and soft-delete recovery.'),
    ('', 12, 'Exercise Phase Management',
     'Structure exercises into named phases (Initial Response, Sustained Operations, Recovery) with inject assignment, phase-grouped MSEL views, and drag-and-drop reordering.'),
    ('', 13, 'Exercise Configuration Wizard',
     'Guided setup with visual progress dashboard covering basic info, participants, time zone, clock mode, objectives, phases, and MSEL. Prevents activation until required items are complete.'),
    ('', 14, 'Clock Mode Selection',
     'Choose between Clock-Driven mode (real-time inject delivery for FE/FSE) or Facilitator-Paced mode (manual sequence advancement for TTX).'),
    ('', 15, 'Participant Assignment',
     'Assign users to HSEEP roles per exercise with role-specific permissions. A user can hold different roles across different exercises.'),
    ('', 16, 'Bulk Participant Import',
     'Upload participant lists via CSV or Excel with preview/validation, handling of existing members, new users, and automatic organization membership management.'),
    ('', 17, 'Inject Approval Workflow',
     'Three-tier configurable approval (organization, exercise, inject level) with submit/approve/reject actions, batch approval, approval queue, Go-Live gate, and configurable permissions including self-approval policies.'),
    ('', 18, 'Excel Import',
     'Guided wizard for importing injects from Excel/CSV files with column mapping, synonym matching for delivery methods, validation before finalization, and create-or-update by inject number.'),
    ('', 19, 'Excel Export',
     'Export MSEL to formatted Excel workbook with round-trip column compatibility, optional metadata/objectives/phases sheets, and blank template download for standardized data entry.'),
    ('', 20, 'MSEL Duplication',
     'Duplicate an existing exercise and its complete MSEL for reuse in recurring exercises (annual drills, quarterly tabletops) with full inject, phase, and objective content preservation.'),
    ('', 21, 'Inject Filtering',
     'Multi-criteria filtering by status, phase, objective, delivery method, and controller assignment. Combinable filters with persistent state within session.'),
    ('', 22, 'Inject Search',
     'Full-text search across inject fields (title, description, sender, recipient, expected action) for rapid content location in large MSELs.'),
    ('', 23, 'Inject Organization',
     'Configurable sorting, grouping by phase/status/objective with collapsible sections, and drag-and-drop reordering with automatic sequence number updates.'),
    ('', 24, 'Setup Progress Dashboard',
     'Visual checklist showing completion status for each configuration area (Basic Info, Roles, Participants, Objectives, Phases, MSEL, Time Zone) with links to incomplete items.'),

    ('Exercise Conduct', 25, 'Real-Time Exercise Clock',
     'Start, pause, and resume exercise timing with elapsed time preservation. Supports independent clock state and exercise status for flexible operational patterns.'),
    ('', 26, 'Clock-Driven Conduct View',
     'Injects organized into Now (ready to fire), Upcoming (within configurable window), and Completed sections. Injects auto-transition to Ready when clock reaches their scheduled delivery time.'),
    ('', 27, 'Facilitator-Paced Conduct View',
     'Sequential inject presentation without time constraints for TTX exercises. Facilitator controls pace with skip, reorder, and revisit capabilities.'),
    ('', 28, 'Story Time Display',
     'Shows scenario narrative time alongside wall clock during exercises using compressed timelines, keeping participants oriented in the story.'),
    ('', 29, 'Inject Firing',
     'Controllers deliver injects with wall clock and scenario time recording, real-time notification to all connected users, and automatic evaluator alerting.'),
    ('', 30, 'Fire Confirmation Dialog',
     'Configurable confirmation dialog for critical injects to prevent accidental delivery during fast-paced conduct.'),
    ('', 31, 'Inject Status Management',
     'Track inject lifecycle through Pending, Ready, Released, Complete, Skipped, and Deferred statuses with required skip reasons and full audit trail.'),
    ('', 32, 'Independent Clock and Status',
     'Exercise status (Active/Paused) and clock state (Running/Paused) operate independently, supporting administrative pauses, clock-only pauses, and combined pauses.'),
    ('', 33, 'Real-Time Notifications',
     'Toast alerts and persistent notification bell with priority-based display (High/Medium/Low), role-targeted delivery, and database-backed persistence across sessions.'),
    ('', 34, 'Auto-Fire with Confirmation',
     'Injects automatically fire at scheduled time with configurable confirmation dialog, reducing Controller workload during high-tempo exercises.'),

    ('Offline Capability & Connectivity', 35, 'Offline Detection',
     'Continuous connectivity monitoring via browser Navigator.onLine and SignalR connection state with clear visual indicators of connection status.'),
    ('', 36, 'Local Data Cache',
     'All active exercise data cached in IndexedDB for full read access to MSEL, inject details, and exercise state regardless of connectivity.'),
    ('', 37, 'Offline Action Queue',
     'User actions (firing injects, recording observations, updating status) queued in persistent FIFO queue in IndexedDB with optimistic UI updates during offline periods.'),
    ('', 38, 'Sync on Reconnect',
     'Queued actions processed sequentially against the server with conflict detection upon connectivity restoration. Last-write-wins for most operations; first-write-wins for inject firing.'),
    ('', 39, 'Zero Data Loss Guarantee',
     'Architecture ensures no user action is lost during offline periods. Queued operations survive browser crashes and are reconciled on reconnect with user notification of conflicts.'),
    ('', 40, 'Real-Time Multi-User Sync',
     'Sub-second synchronization of inject status changes, clock events, observations, and participant activity across all connected users via Azure SignalR Service.'),

    ('Observation & Evaluation', 41, 'Ad-Hoc Observation Capture',
     'Quick-entry interface for real-time observation during conduct with type classification (Strength, Area for Improvement, Neutral), optional P/S/M/U rating, and dual timestamps.'),
    ('', 42, 'Observation-Inject Linkage',
     'Many-to-many association between observations and injects, enabling post-exercise analysis of which injects generated which performance observations.'),
    ('', 43, 'Observation-Objective Linkage',
     'Many-to-many association between observations and exercise objectives for coverage analysis and AAR preparation.'),
    ('', 44, 'Structured EEG Entry',
     'Record structured assessments against specific Critical Tasks with mandatory P/S/M/U rating, optional inject linkage, and evaluator identification.'),
    ('', 45, 'EEG Coverage Dashboard',
     'Real-time visualization of Critical Task evaluation coverage, P/S/M/U rating distribution by Capability Target, and identification of unevaluated tasks.'),
    ('', 46, 'Observation Filtering',
     'Filter observations by type, rating, objective, evaluator, and time range for efficient review and AAR preparation.'),
    ('', 47, 'Offline Observation Support',
     'Observations persist through connectivity loss in IndexedDB and sync automatically on reconnect with conflict resolution and duplicate prevention.'),

    ('Field Operations', 48, 'Photo Capture and Attachment',
     'Field participants capture and attach photos to observations with optional annotations. Client-side compression for bandwidth efficiency with offline queue for upload.'),
    ('', 49, 'Voice-to-Text Observation Input',
     'Browser-based speech recognition for hands-free observation entry in the field, with full text editing after capture. No cloud dependency.'),
    ('', 50, 'GPS Location Tracking',
     'Opt-in participant location sharing with real-time Director map view, automatic geo-stamping of observations and inject firings, and automatic disable when exercise clock stops.'),
    ('', 51, 'Director Situational Awareness',
     'Real-time map of field team positions, incoming observation feed, coverage gap identification, and safety alert visibility for the Exercise Director.'),
    ('', 52, 'Safety-Flagged Observations',
     'Immediate flagging of real-world safety concerns with photo evidence and location data for Safety Officer and Exercise Director visibility.'),

    ('After-Action Review & Reporting', 53, 'Review Mode',
     'Dedicated post-conduct view with phase-grouped timeline, inject outcome summaries with time variance, observation review panel, and exercise statistics optimized for AAR discussion.'),
    ('', 54, 'Exercise Progress Metrics',
     'Real-time inject delivery statistics (fired, skipped, pending, deferred), on-time delivery rate, and timing variance during and after conduct.'),
    ('', 55, 'Observation Summary Metrics',
     'Observation counts, P/S/M/U rating distribution, coverage rate by objective, and evaluator activity analysis.'),
    ('', 56, 'Timeline Analysis',
     'Planned vs. actual duration, time per phase, pause count and duration, and inject pacing analysis for post-exercise evaluation.'),
    ('', 57, 'Core Capability Performance',
     'Performance metrics organized by FEMA Core Capability for alignment with national preparedness goals.'),
    ('', 58, 'Controller and Evaluator Metrics',
     'Controller workload distribution and timing analysis. Evaluator observation coverage matrix and rating consistency.'),
    ('', 59, 'Organization Performance Trends',
     'Cross-exercise P/S/M/U trends, capability performance tracking, improvement area identification, and exercise program effectiveness analysis.'),
    ('', 60, 'Comparative Analysis',
     'Side-by-side metric comparison between exercises for improvement tracking and benchmarking.'),
    ('', 61, 'EEG-Based AAR Export',
     'Export findings organized by Capability, Capability Target, Critical Task, and Observations in HSEEP AAR/IP format.'),
    ('', 62, 'Metrics Export',
     'Export all metrics to PDF, Excel, and image formats for inclusion in reports and briefings.'),
    ('', 63, 'MSEL and Observation Excel Export',
     'Export exercise data for stakeholders without platform access, with formatted workbooks including metadata sheets.'),
    ('', 64, 'Custom Metrics Dashboard',
     'User-configurable metric views with selectable widgets, date ranges, and filter criteria for tailored analysis.'),

    ('Multi-Tenancy & Organization Management', 65, 'Organization-Based Multi-Tenancy',
     'Organization as the primary security and data isolation boundary. All exercise data scoped to organizations with validated query-level isolation and write-side interceptors.'),
    ('', 66, 'Three-Tier Role Hierarchy',
     'System roles (Admin, Manager, User), Organization roles (OrgAdmin, OrgManager, OrgUser), and Exercise roles (9 HSEEP roles) with per-context permissions.'),
    ('', 67, 'Multi-Organization Membership',
     'Users belong to multiple organizations with different roles in each, supporting consultants and contractors who work across agencies. Seamless organization switcher.'),
    ('', 68, 'Organization Lifecycle',
     'Active, Archived, and Inactive states with appropriate access controls and data retention for each state.'),
    ('', 69, 'User Invitation System',
     'Invite users via email or shareable organization codes for self-service onboarding with configurable code expiration.'),
    ('', 70, 'Agency List Management',
     'Define participating agencies (Fire, EMS, Police, Public Health) per organization and assign them to participants, injects, and observations.'),
    ('', 71, 'Capability Library Selection',
     'Choose from pre-built capability frameworks (FEMA Core Capabilities, NATO, NIST CSF, ISO 22301) or define custom capabilities per organization.'),
    ('', 72, 'Organization Settings',
     'Configurable defaults for session timeout, auto-save interval, exercise templates, branding, and core capability lists at the organizational level.'),

    ('Authentication & Security', 73, 'Local Credential Authentication',
     'Email and password authentication with JWT access tokens (15-minute, memory-stored) and refresh tokens (4-hour, HttpOnly cookies).'),
    ('', 74, 'Azure Entra SSO Integration',
     'Enterprise Single Sign-On via Azure Active Directory with automatic account linking by email for organizations using Microsoft identity.'),
    ('', 75, 'Exercise-Scoped Authorization',
     'Per-exercise role assignments with role inheritance from global defaults. Server-side validation on every API request.'),
    ('', 76, 'Token Management',
     'Silent automatic refresh, 4+ hour session duration for exercise conduct, multi-tab session sharing, and graceful expiration handling.'),
    ('', 77, 'First-User Bootstrap',
     'First registered user automatically becomes Administrator, eliminating deployment configuration for initial setup.'),
    ('', 78, 'User Management',
     'Admin tools for user listing, creation, editing, deactivation, role assignment, and inline user creation during exercise setup.'),

    ('Platform Experience', 79, 'Role-Aware Home Dashboard',
     'Personalized landing page with role-appropriate welcome message, quick actions based on permissions, and recent exercise list with status indicators.'),
    ('', 80, 'Context-Aware Navigation',
     'Sidebar transforms when entering exercise context, showing conduct-relevant options (MSEL, Inject Queue, Observations) with role-based menu visibility.'),
    ('', 81, 'Exercise Clock in Header',
     'Persistent clock display showing elapsed time, clock state, and story time during exercise conduct, visible from any page within the exercise.'),
    ('', 82, 'Extended Session Support',
     'Configurable session timeout (default 4 hours) with warning before expiration, designed for multi-hour exercise conduct sessions.'),
    ('', 83, 'Auto-Save',
     'Automatic data persistence on field blur and configurable intervals, preventing data loss from accidental navigation or browser crashes.'),
    ('', 84, 'Keyboard Navigation',
     'Comprehensive keyboard shortcuts for rapid inject management and primary actions, reducing mouse dependency during fast-paced conduct.'),
    ('', 85, 'Responsive Design',
     'Fully functional on desktop and tablet devices (1024px+) with touch-optimized interactions for field use.'),
    ('', 86, 'Three-Tier Settings Model',
     'User preferences (persistent across exercises), exercise settings (per-exercise), and organization defaults (inherited unless overridden) with auto-save and reset-to-default.'),
    ('', 87, 'My Assignments Dashboard',
     'Personalized view of exercise assignments grouped by Active, Upcoming, and Completed with role-based landing routes.'),
    ('', 88, 'Semantic Versioning and Release Notes',
     'Automated version tracking with in-app What\'s New notifications and release history for transparency on platform updates.'),

    ('Technical Infrastructure', 89, 'Always-Warm API',
     'REST API hosted on Azure App Service with no cold starts, ensuring instant response times critical for exercise conduct. Serverless functions used only for background maintenance tasks.'),
    ('', 90, 'Real-Time Infrastructure',
     'Azure SignalR Service providing sub-second event propagation for inject status changes, clock events, observations, and participant activity.'),
    ('', 91, 'Cloud-Native Deployment',
     'Microsoft Azure hosting with CI/CD via GitHub Actions, infrastructure as code (Bicep templates), and environment-specific configuration for development, staging, and production.'),
    ('', 92, 'Modular Feature Architecture',
     'Self-contained feature modules with clear separation between business logic and web infrastructure, supporting independent development, testing, and deployment.'),
]

# -- Headers --
headers = ['#', 'Category', 'Feature', 'Description']
col_widths = [5, 30, 30, 80]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# -- Data Rows --
current_category = ''
row_num = 2
for cat, num, feature, desc in features:
    if cat:
        current_category = cat

    ws.cell(row=row_num, column=1, value=num).font = normal_font
    ws.cell(row=row_num, column=1).alignment = center_alignment
    ws.cell(row=row_num, column=1).border = thin_border

    cat_cell = ws.cell(row=row_num, column=2, value=current_category if cat else '')
    cat_cell.font = category_font if cat else normal_font
    cat_cell.alignment = wrap_alignment
    cat_cell.border = thin_border
    if cat:
        cat_cell.fill = category_fill

    feat_cell = ws.cell(row=row_num, column=3, value=feature)
    feat_cell.font = Font(name='Calibri', size=10, bold=True)
    feat_cell.alignment = wrap_alignment
    feat_cell.border = thin_border

    desc_cell = ws.cell(row=row_num, column=4, value=desc)
    desc_cell.font = normal_font
    desc_cell.alignment = wrap_alignment
    desc_cell.border = thin_border

    # Alternating row color (within category groups)
    if not cat and row_num % 2 == 0:
        for c in range(1, 5):
            if not ws.cell(row=row_num, column=c).fill.start_color.rgb or ws.cell(row=row_num, column=c).fill.start_color.rgb == '00000000':
                ws.cell(row=row_num, column=c).fill = alt_fill

    row_num += 1

# -- Freeze panes --
ws.freeze_panes = 'A2'

# -- Auto-filter --
ws.auto_filter.ref = f'A1:D{row_num - 1}'

# -- Summary sheet --
ws2 = wb.create_sheet('Summary')
ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 12

summary_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
summary_data = [
    ('Category', 'Count'),
    ('HSEEP Compliance & Framework Support', 7),
    ('Exercise Design & Planning', 17),
    ('Exercise Conduct', 10),
    ('Offline Capability & Connectivity', 6),
    ('Observation & Evaluation', 7),
    ('Field Operations', 5),
    ('After-Action Review & Reporting', 12),
    ('Multi-Tenancy & Organization Management', 8),
    ('Authentication & Security', 6),
    ('Platform Experience', 10),
    ('Technical Infrastructure', 4),
    ('TOTAL', 92),
]

for r_idx, (cat, count) in enumerate(summary_data, 1):
    cat_cell = ws2.cell(row=r_idx, column=1, value=cat)
    count_cell = ws2.cell(row=r_idx, column=2, value=count)
    cat_cell.border = thin_border
    count_cell.border = thin_border
    count_cell.alignment = Alignment(horizontal='center')

    if r_idx == 1:
        cat_cell.font = summary_header_font
        cat_cell.fill = header_fill
        count_cell.font = summary_header_font
        count_cell.fill = header_fill
    elif r_idx == len(summary_data):
        cat_cell.font = Font(name='Calibri', size=11, bold=True)
        count_cell.font = Font(name='Calibri', size=11, bold=True)
        cat_cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        count_cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    else:
        cat_cell.font = Font(name='Calibri', size=10)
        count_cell.font = Font(name='Calibri', size=10)

output_path = r'C:\Code\dynamis\cadence\docs\bd\Cadence_Feature_Matrix.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
